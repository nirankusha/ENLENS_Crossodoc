#!/usr/bin/env python3
import argparse
from openpyxl import load_workbook
import pandas as pd
import re
import colorsys
import os

# --- Color conversion functions ---
def rgb_to_hsv(rgb_hex):
    if rgb_hex is None:
        return None
    if len(rgb_hex) == 8:
        rgb_hex = rgb_hex[2:]
    r = int(rgb_hex[0:2], 16) / 255
    g = int(rgb_hex[2:4], 16) / 255
    b = int(rgb_hex[4:6], 16) / 255
    h, s, v = colorsys.rgb_to_hsv(r, g, b)
    return round(h * 360, 1), round(s, 3), round(v, 3)

def rgb_to_signed_score(rgb_hex):
    if rgb_hex is None:
        return 0
    if len(rgb_hex) == 8:
        rgb_hex = rgb_hex[2:]
    r = int(rgb_hex[0:2], 16) / 255
    g = int(rgb_hex[2:4], 16) / 255
    b = int(rgb_hex[4:6], 16) / 255
    h, s, v = colorsys.rgb_to_hsv(r, g, b)
    hue_deg = h * 360
    intensity = s * v
    if 90 <= hue_deg <= 150:    # green
        return intensity
    elif hue_deg <= 15 or hue_deg >= 345:  # red
        return -intensity
    else:
        return 0

def theme_to_rgb(theme_index):
    theme_map = {
        9: 'FF00FF00',   # green
        10: 'FFFF0000',  # red
        0: None,
    }
    return theme_map.get(theme_index, None)

def signed_score_to_likert(score, n_levels=5):
    if score == 0:
        return 0
    abs_score = min(abs(score), 1.0)
    level = max(1, round(abs_score * n_levels))
    return level if score > 0 else -level

# --- Main extraction function ---
def extract_third_sheet_multiheader_and_urls(
    xlsx_path: str,
    header_rows: tuple = (3, 5),
    annotate_shaded: bool = True,
    n_levels: int = 5
) -> pd.DataFrame:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[2]
    all_rows = []

    for row in ws.iter_rows(values_only=False):
        row_out = []
        for cell in row:
            val = cell.value
            score = 0
            fill = cell.fill

            if fill and fill.patternType:
                try:
                    t = getattr(fill.fgColor, 'type', None)
                    if t == 'rgb' and fill.fgColor.rgb:
                        score = rgb_to_signed_score(fill.fgColor.rgb.upper())
                    elif t == 'theme' and fill.fgColor.theme is not None:
                        rgb = theme_to_rgb(fill.fgColor.theme)
                        score = rgb_to_signed_score(rgb)
                except AttributeError:
                    pass

            likert = signed_score_to_likert(score, n_levels=n_levels)
            if val is None:
                cell_out = likert
            else:
                cell_out = f"{val}|{likert}" if annotate_shaded else val

            row_out.append(cell_out)
        all_rows.append(row_out)

    df = pd.DataFrame(all_rows)

    # --- MultiIndex header ---
    idx1, idx2 = header_rows[0] - 1, header_rows[1] - 1
    lvl1 = df.iloc[idx1].tolist()
    lvl2 = df.iloc[idx2].tolist()
    filled = []
    last = None
    for v in lvl1:
        if v is not None:
            filled.append(v)
            last = v
        else:
            filled.append(last)
    lvl1 = filled

    tuples = [
        (l1 if l1 is not None else f"Unnamed1_{i}",
         l2 if l2 is not None else f"Unnamed2_{i}")
        for i, (l1, l2) in enumerate(zip(lvl1, lvl2))
    ]
    df.columns = pd.MultiIndex.from_tuples(tuples)
    df = df.drop(index=range(0, idx2 + 1)).reset_index(drop=True)

    # --- Extract URLs ---
    url_pattern = r'(https?://[^\s]+)'
    ref_level2 = [c for c in df.columns.get_level_values(1) if "Reference" in str(c)][0]
    ref_ser = df.xs(ref_level2, axis=1, level=1).squeeze()
    df[('URLs', 'first_url')] = ref_ser.str.extract(url_pattern, expand=False)
    df[('URLs', 'all_urls')] = ref_ser.apply(lambda txt: re.findall(url_pattern, str(txt)))

    return df

# --- Flatten MultiIndex into a "flat" DataFrame ---
# --- Flatten MultiIndex into a "flat" DataFrame with URLs ---
def flatten_df(df: pd.DataFrame) -> pd.DataFrame:
    import re

    # Flatten MultiIndex
    df.columns = ["|".join([str(c) for c in col if c not in (None, "")])
                  for col in df.columns.to_flat_index()]

    flat_df = pd.DataFrame({
        "Goal": df.get("Input data|Goal"),
        "Target": df.get("Input data|Target"),
        "Target description": df.get("Input data|Target description"),
        "Dimension of Temporality": df.get("Input data|Dimension of Temporality"),
        "Reciprocal Interdependence": df.get("Input data|Reciprocal Interdependence"),
        "Justification": df.get("Green Hydrogen Value Chain Justification|COLORED_EMPTY|Justification"),
        "Reference": df.get("Green Hydrogen Value Chain Justification|COLORED_EMPTY|Reference"),
    })

    # Forward-fill Goal across Targets
    flat_df["Goal"] = flat_df["Goal"].ffill()

    # Replace empty strings with None
    flat_df = flat_df.replace(r'^\s*$', None, regex=True)

    # Drop rows without Target
    flat_df = flat_df.dropna(subset=["Target"], how="all")

    # --- URL extraction ---
    url_pattern = r'(https?://[^\s]+)'
    flat_df['extracted_url'] = flat_df['Reference'].str.extract(url_pattern, expand=False)
    flat_df['all_urls'] = flat_df['Reference'].apply(lambda x: re.findall(url_pattern, str(x)))

    return flat_df

# --- Command-line interface ---
def main():
    import argparse
    parser = argparse.ArgumentParser(description="Extract Excel sheet with color-coded ordinal values and flatten.")
    parser.add_argument("xlsx_path", type=str, help="Path to Excel file")
    parser.add_argument("output_dir", type=str, help="Directory to save CSV output")
    parser.add_argument("--header_rows", type=int, nargs=2, default=[3,5], help="Header rows for MultiIndex")
    parser.add_argument("--n_levels", type=int, default=5, help="Number of Likert levels")
    parser.add_argument("--annotate_shaded", action="store_true", help="Append ordinal value to cell text")
    args = parser.parse_args()

    df = extract_third_sheet_multiheader_and_urls(
        args.xlsx_path,
        header_rows=tuple(args.header_rows),
        annotate_shaded=args.annotate_shaded,
        n_levels=args.n_levels
    )

    flat_df = flatten_df(df)

    # Save CSV
    os.makedirs(args.output_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(args.xlsx_path))[0]
    out_path = os.path.join(args.output_dir, f"{base_name}_flattened.csv")
    flat_df.to_csv(out_path, index=False)
    print(f"Saved flattened CSV to: {out_path}")
    print(flat_df.head(10))

if __name__ == "__main__":
    main()
